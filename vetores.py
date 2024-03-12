from openpyxl import load_workbook

# Carregue o arquivo Excel
wb = load_workbook(filename='planilha.xlsx')

# Selecione a planilha
sheet = wb.active

# Inicialize listas / vetores vazios para armazenar os dados
nomes = []
idades = []
pesos = []

# Itere sobre as colunas e armazene os valores nas listas correspondentes
for row in sheet.iter_rows(min_row=2, values_only=True):
    nomes.append(row[0])   # A primeira coluna contém nomes
    idades.append(row[1])  # A segunda coluna contém idades
    pesos.append(row[2])   # A terceira coluna contém pesos

# Exemplo de acesso aos dados
print("Nomes = ", nomes)  # Lista de nomes
print("Idades = ", idades)  # Lista de idades
print("PEsos = ", pesos)  # Lista de pesos

# iter_rows(min_row=2, values_only=True) itera sobre as linhas do Excel, começando da segunda linha (para evitar o cabeçalho);
# values_only = True garante que obtenhamos apenas os valores das células, não objetos de célula;
# row[0], row[1] e row[2] acessam os valores das colunas "Nome", "Idade" e "Peso", respectivamente, em cada linha e os adicionam às listas correspondentes.
